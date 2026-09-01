#!/usr/bin/env python3
"""Turn a raw VTS export into the report you send your landlord client.

VTS downloads `leasing-activity-MM-DD-YY.xlsx` containing two sheets: an
"Overview" tab with building stats, and the deals sheet. All this does is delete
the Overview tab and give the file a sensible name. The deals sheet is left
completely untouched, because its formatting is already what gets sent.

**The destination is optional.** Plenty of people never keep leasing reports —
they export, email the landlord, and delete. With no destination the finished
file lands next to the export (normally Downloads), named from the property.

Where a destination folder IS given and already holds past reports, the naming
convention is copied from the newest one rather than invented, since it differs
per property (street address for some, property name for others).

    python finalize_report.py <raw_export.xlsx> [dest_folder]
        [--out DIR] [--name "Property Name"] [--date M.D.YY] [--dry-run]
"""
import datetime as dt
import os
import pathlib
import re
import shutil
import sys

import os as _os
import sys as _sys
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)),
                                  "..", "..", "..", "lib"))
from vts_env import ensure_deps
from vts_errors import check_folder, friendly_move, friendly_open, friendly_save
ensure_deps()          # re-execs under the plugin venv if openpyxl is missing

import openpyxl


def expand(p):
    """Windows shells do not expand ~, so every path arg goes through here."""
    return str(pathlib.Path(p).expanduser())


NAME_RE = re.compile(
    r"^(?P<prefix>.*?)(?P<m>\d{1,2})(?P<sep>[.\-])(?P<d>\d{1,2})(?P=sep)(?P<y>\d{2,4})\.xlsx$",
    re.IGNORECASE,
)


def existing_reports(folder):
    out = []
    for f in os.listdir(folder):
        if f.startswith("~$") or not f.lower().endswith(".xlsx"):
            continue
        # Raw VTS downloads are not reports. Modelling a name on one produces
        # "leasing-activity-9-1-26.xlsx", which is not something you send a landlord.
        if f.lower().startswith("leasing-activity"):
            continue
        p = os.path.join(folder, f)
        if os.path.isfile(p):
            out.append((os.path.getmtime(p), f))
    return [f for _, f in sorted(out, reverse=True)]


ILLEGAL = re.compile(r'[<>:"/\\|?*]')


def property_title(wb):
    """The property name, read from the deals sheet itself.

    Used to name the output when there's no prior report to copy from. Reading it
    from the workbook means the name doesn't depend on how anyone files things.
    """
    for name in wb.sheetnames:
        if name.lower().startswith("overview"):
            continue
        ws = wb[name]
        for r in range(1, 6):
            for c in range(1, 10):
                v = ws.cell(row=r, column=c).value
                if v and str(v).strip():
                    return str(v).strip()
        return name
    return "Property"


def target_name(out_dir, today, fallback_name, match_convention=True):
    """Name the finished report.

    An explicit name always wins. Otherwise, in a folder the user actually chose,
    copy the newest prior report's convention. In the implicit fallback location
    (next to the download) don't copy anything — whatever else is sitting in
    Downloads has nothing to do with this landlord's report.
    """
    stamp_dot = f"{today.month}.{today.day}.{today.strftime('%y')}"
    for f in (existing_reports(out_dir) if match_convention else []):
        m = NAME_RE.match(f)
        if m:
            sep = m.group("sep")
            stamp = f"{today.month}{sep}{today.day}{sep}{today.strftime('%y')}"
            return f"{m.group('prefix')}{stamp}.xlsx", f
    base = ILLEGAL.sub("", str(fallback_name or "Property")).strip() or "Property"
    return f"{base} Leasing Update {stamp_dot}.xlsx"[:200], None


def finalize(raw_path, folder=None, date=None, dry_run=False, name=None):
    # No destination given: put it beside the export, i.e. wherever the browser
    # downloaded it. Someone who doesn't archive reports still gets a named file.
    chose_folder = bool(folder)
    if folder:
        folder = check_folder(folder, "destination folder")
    else:
        folder = str(pathlib.Path(raw_path).expanduser().resolve().parent)
    today = dt.date.today()
    if date:
        m, d, y = re.split(r"[.\-/]", date)
        y = int(y) + 2000 if len(y) == 2 else int(y)
        today = dt.date(y, int(m), int(d))

    wb = friendly_open(raw_path, "VTS export")
    fname, modeled_on = target_name(folder, today, name or property_title(wb),
                                    match_convention=chose_folder and not name)
    dest = os.path.join(folder, fname)
    removed = [s for s in wb.sheetnames if s.lower().startswith("overview")]

    print(f"source:     {raw_path}")
    print(f"sheets:     {wb.sheetnames}")
    print(f"removing:   {removed or 'nothing (no Overview tab found)'}")
    print(f"naming:     {fname}" + (f"   (matched: {modeled_on})" if modeled_on
                                    else "   (named from the property; no prior report here to copy)"))
    print(f"dest:       {dest}")

    if dry_run:
        print("\n[dry run — nothing written]")
        return dest

    if os.path.exists(dest):
        backup = dest.replace(".xlsx", f".superseded-{dt.datetime.now():%H%M%S}.xlsx")
        friendly_move(dest, backup)
        print(f"existing file moved aside -> {os.path.basename(backup)}")

    for s in removed:
        del wb[s]
    friendly_save(wb, dest)
    print(f"\nwrote {dest}")
    return dest


def _opt(flag):
    """Read --flag=value or --flag value."""
    for i, a in enumerate(sys.argv[1:], start=1):
        if a == flag and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
        if a.startswith(flag + "="):
            return a.split("=", 1)[1]
    return None


def main():
    flags = {"--date", "--out", "--name"}
    args, skip = [], False
    for a in sys.argv[1:]:
        if skip:
            skip = False
            continue
        if a in flags:
            skip = True
            continue
        if a.startswith("--"):
            continue
        args.append(a)

    if not args:
        print(__doc__)
        sys.exit(1)

    out = _opt("--out")
    folder = out or (args[1] if len(args) > 1 else None)
    finalize(expand(args[0]),
             expand(folder) if folder else None,
             date=_opt("--date"),
             dry_run="--dry-run" in sys.argv,
             name=_opt("--name"))


if __name__ == "__main__":
    main()
